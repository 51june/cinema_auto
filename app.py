import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io
import base64
import json
import requests
import re

st.set_page_config(
    page_title="影院票房与座位自动统计工具",
    page_icon="🎬",
    layout="centered",
    initial_sidebar_state="collapsed"
)

st.markdown("""
    <style>
    .main { background-color: #fafbfa; }
    .stButton>button { width: 100%; background-color: #1F4E78; color: white; border-radius: 8px; padding: 10px; font-weight: bold; }
    .stButton>button:hover { background-color: #153654; color: white; }
    .upload-box { border: 2px dashed #1F4E78; padding: 20px; border-radius: 10px; text-align: center; background-color: #ffffff; }
    h1 { color: #1F4E78; font-family: 'Microsoft YaHei', sans-serif; }
    </style>
""", unsafe_allow_html=True)

st.title("🎬 影院座位截图自动录入系统")
st.write("上传购票界面的座位截图，系统将自动识别并更新至 Excel 表格中。自动识别同场次并覆盖旧数据。")

# ================= 侧边栏：配置与历史数据导入 =================
st.sidebar.header("⚙️ 配置中心")
api_key = st.sidebar.text_input("请输入大模型 API Key", type="password")
model_provider = st.sidebar.selectbox("选择AI模型供应商", ["智谱清言 GLM-4V (国内直连推荐)", "Gemini (需科学上网)"])

st.sidebar.markdown("---")
st.sidebar.header("📂 历史数据记忆")
uploaded_excel = st.sidebar.file_uploader("上传您上次下载的 Excel 统计表，系统将在此基础上更新覆盖", type=["xlsx"])

# 初始化或加载表格数据
if uploaded_excel is not None:
    try:
        # 读取用户上传的历史Excel
        df_uploaded = pd.read_excel(uploaded_excel)
        # 将导出的复杂表头还原为系统内部表头
        df_uploaded = df_uploaded.rename(columns={
            "影院名称(预售开启后更新影院列表)": "影院名称",
            "最后更新时间(精确到分)": "最后更新时间"
        })
        # 移除公式列，后续重新生成
        if '占比' in df_uploaded.columns:
            df_uploaded = df_uploaded.drop(columns=['占比'])
        # 填充空值
        df_uploaded = df_uploaded.fillna("")
        st.session_state['excel_data'] = df_uploaded
        st.sidebar.success("✅ 历史表格加载成功！系统已恢复记忆，接下来的识别将覆盖表中的原数据。")
    except Exception as e:
        st.sidebar.error(f"读取 Excel 失败，请检查文件: {e}")
elif 'excel_data' not in st.session_state:
    # 如果没传表，就用空白模板
    df_init = pd.DataFrame(columns=["影院名称", "日期", "时间档", "总座位数", "已售", "最后更新时间"])
    st.session_state['excel_data'] = df_init


uploaded_file = st.file_uploader("📸 拍照或选择手机相册中的截图", type=["jpg", "jpeg", "png"])

def encode_image(uploaded_file):
    return base64.b64encode(uploaded_file.read()).decode('utf-8')

def analyze_image_with_ai(image_base64, api_key, provider):
    prompt = """
    你是一个专业的影院座位数据分析和极其严谨的视觉计数专家。
    AI系统往往会低估或高估密集的座位网格数量。你必须使用最高级别的精度重新清点！必须把每个可见的座位格子当成离散实体一个一个数！

    请提取以下信息：
    1. 影院名称（例如：UME影城（上海新天地店）规范化为“UME影城 新天地”）
    2. 观影日期（严格提取纯日期，例如从“下周五 6月19日”中只提取“6月19日”，过滤掉“今天/明天/周几”等修饰词）
    3. 时间档：请提取电影的开始时间（例如：13:50）。严格过滤散场时间。
    4. 截图左上角的手机系统时间（作为最后更新时间）。
    5. 总座位数：必须按排从左到右极其精确地计数！
       - 极其重要防错指令：绝不能把底部‘推荐座位’或‘图例’区域的文字数字（如 1人, 2人, 3人, 4人, 5人）当成座位数加进去！
       - 无论图片是哪个影厅，必须如实清点画面中实际供人乘坐的矩形方格数量，不要带有任何预设数字。
    6. 已售座位数：仔细清点所有变红或带有猫头/人物图像的红色已售方格数量。如果不确定，或者只看到空座，必须输出 0！

    请严格以 JSON 格式输出，不要包含任何 Markdown 标记或多余文字。即使找不到某个信息，也要输出该字段并填空字符串，绝不能遗漏字段！
    {
        "cinema_name": "示例影城",
        "date": "1月1日",
        "time_slot": "12:00",
        "total_seats": 0,
        "sold_seats": 0,
        "update_time": "12:00"
    }
    """
    
    clean_key = re.sub(r'[^a-zA-Z0-9._-]', '', api_key)
    
    if provider == "智谱清言 GLM-4V (国内直连推荐)":
        url = base64.b64decode("aHR0cHM6Ly9vcGVuLmJpZ21vZGVsLmNuL2FwaS9wYWFzL3Y0L2NoYXQvY29tcGxldGlvbnM=").decode("utf-8")
        headers = {
            "Authorization": f"Bearer {clean_key}",
            "Content-Type": "application/json"
        }
        payload = {
            "model": "glm-4v",
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_base64}"}}
                    ]
                }
            ]
        }
        try:
            response = requests.post(url, headers=headers, json=payload, timeout=30)
            res_json = response.json()
            if 'error' in res_json:
                st.error(f"API 错误: {res_json['error']['message']}")
                return None
            text_response = res_json['choices'][0]['message']['content']
            text_response = text_response.replace("```json", "").replace("```", "").strip()
            return json.loads(text_response)
        except Exception as e:
            st.error(f"AI 识别解析出错啦: {str(e)}")
            return None
            
    elif provider == "Gemini (需科学上网)":
        base_url = base64.b64decode("aHR0cHM6Ly9nZW5lcmF0aXZlbGFuZ3VhZ2UuZ29vZ2xlYXBpcy5jb20vdjFiZXRhL21vZGVscy9nZW1pbmktMS41LWZsYXNoOmdlbmVyYXRlQ29udGVudD9rZXk9").decode("utf-8")
        url = base_url + clean_key
        headers = {'Content-Type': 'application/json'}
        payload = {"contents": [{"parts": [{"text": prompt}, {"inlineData": {"mimeType": "image/jpeg", "data": image_base64}}]}]}
        try:
            response = requests.post(url, headers=headers, json=payload, timeout=30)
            res_json = response.json()
            text_response = res_json['candidates'][0]['content']['parts'][0]['text']
            text_response = text_response.replace("```json", "").replace("```", "").strip()
            return json.loads(text_response)
        except Exception as e:
            st.error(f"AI 识别解析出错啦: {str(e)}")
            return None

# ================= 规范化匹配字段，防止 AI 输出细微格式误差导致重复新建行 =================
def standardize_date(date_str):
    nums = re.findall(r'\d+', str(date_str))
    if len(nums) >= 2:
        return f"{int(nums[0])}月{int(nums[1])}日" # 强制把 06月19日 转换成 6月19日
    return str(date_str).strip()

def standardize_time(time_str):
    time_match = re.search(r'\d{1,2}:\d{2}', str(time_str))
    if time_match:
        return time_match.group()
    return str(time_str).strip()

if uploaded_file is not None:
    st.image(uploaded_file, caption='已上传的截图 preview', use_container_width=True)
    
    if st.button("🚀 开始自动分析并录入表格"):
        if not api_key:
            st.warning("⚠️ 无法真实分析：请输入 API Key 后再点击本按钮。")
            result = None
        else:
            with st.spinner("AI 正在严谨地计数座位中，请稍候..."):
                img_b64 = encode_image(uploaded_file)
                result = analyze_image_with_ai(img_b64, api_key, model_provider)
        
        if result:
            raw_cinema = str(result.get("cinema_name", "未知影院")).strip()
            raw_date = result.get("date", "未知日期")
            raw_time = result.get("time_slot", "")
            
            # 净化字段格式
            cinema_name = raw_cinema
            date_str = standardize_date(raw_date)
            time_slot = standardize_time(raw_time)
            
            total_seats = result.get("total_seats", 0)
            sold_seats = result.get("sold_seats", 0)
            update_time = str(result.get("update_time", "")).strip()
            
            st.success("🎉 数据处理成功！")
            
            df = st.session_state['excel_data'].copy()
            
            # 确保原表的数据也经过 strip 处理再比对，实现 100% 完美匹配
            exact_match = (df['影院名称'].astype(str).str.strip() == cinema_name) & \
                          (df['日期'].astype(str).str.strip() == date_str) & \
                          (df['时间档'].astype(str).str.strip() == time_slot)
            
            if exact_match.any():
                idx = df[exact_match].index[0]
                df.loc[idx, '总座位数'] = total_seats
                df.loc[idx, '已售'] = sold_seats
                df.loc[idx, '最后更新时间'] = update_time
                st.info(f"🔄 检测到相同场次 ({cinema_name} | {date_str} {time_slot})，已自动为您覆盖更新旧数据。")
            else:
                new_row = {
                    "影院名称": cinema_name, 
                    "日期": date_str, 
                    "时间档": time_slot, 
                    "总座位数": total_seats, 
                    "已售": sold_seats, 
                    "最后更新时间": update_time
                }
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                st.success(f"➕ 发现新场次 ({cinema_name} | {date_str} {time_slot})，已自动新增一行。")
                
            st.session_state['excel_data'] = df

st.subheader("📊 当前实时统计报表（数据预览）")
st.dataframe(st.session_state['excel_data'], use_container_width=True)

def convert_df_to_excel(df_data):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "票房与座位统计表"
    ws.views.sheetView[0].showGridLines = True
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Microsoft YaHei", size=10, bold=False, color="000000")
    thin_side = Side(border_style="thin", color="D9D9D9")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    headers = ["影院名称(预售开启后更新影院列表)", "日期", "时间档", "总座位数", "已售", "占比", "最后更新时间(精确到分)"]
    ws.append(headers)
    ws.row_dimensions[1].height = 28
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_all

    for r_idx, row in enumerate(df_data.itertuples(index=False), 2):
        row_vals = [row[0], row[1], row[2], row[3], row[4], "", row[5]]
        ws.append(row_vals)
        ws.row_dimensions[r_idx].height = 22
        
        for c_idx in range(1, 8):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = border_all
            cell.fill = zebra_fill if r_idx % 2 == 0 else white_fill
            
            if c_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_idx in [2, 3, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx in [4, 5]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if cell.value != "": cell.number_format = '#,##0'
            elif c_idx == 6:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '0.0%'
                cell.value = f'=IFERROR(E{r_idx}/D{r_idx}, "")'
                
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['G'].width = 24
    wb.save(output)
    return output.getvalue()

if st.button("📥 下载已更新的专业 Excel 报表"):
    excel_bytes = convert_df_to_excel(st.session_state['excel_data'])
    st.download_button(
        label="点击下载 .xlsx 文件",
        data=excel_bytes,
        file_name="最新影院预售统计表.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
